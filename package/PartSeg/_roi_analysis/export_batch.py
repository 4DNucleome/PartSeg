import json
import logging
import os
import tarfile
import time
import typing
import zipfile
from contextlib import suppress
from pathlib import Path

import requests
from openpyxl import load_workbook
from qtpy.QtCore import QByteArray, Qt
from qtpy.QtGui import QIcon
from qtpy.QtWidgets import (
    QCheckBox,
    QDialog,
    QFormLayout,
    QGridLayout,
    QLabel,
    QLineEdit,
    QProgressBar,
    QPushButton,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QWidget,
)
from superqt import QCollapsible
from superqt.utils import thread_worker

from PartSeg._roi_analysis.partseg_settings import PartSettings
from PartSeg.common_backend.base_settings import IO_SAVE_DIRECTORY
from PartSeg.common_gui.custom_load_dialog import PLoadDialog, SelectDirectoryDialog
from PartSeg.common_gui.custom_save_dialog import PSaveDialog
from PartSeg.common_gui.main_window import OPEN_DIRECTORY
from PartSeg.common_gui.select_multiple_files import IO_BATCH_DIRECTORY
from PartSegCore.io_utils import LoadPlanExcel
from PartSegData import icons_dir

REQUESTS_TIMEOUT = 600

NO_FILES = "No files to export"
MISSING_FILES = "Some files do not exists"


class ExportProjectDialog(QDialog):
    """Export data for zenodo"""

    def __init__(
        self, excel_path: str, base_folder: str, settings: PartSettings, parent: typing.Optional[QWidget] = None
    ) -> None:
        super().__init__(parent=parent)
        self.setWindowTitle("Export batch with data")
        self.settings = settings
        self._all_files_exists = False
        self.info_label = QLabel()
        self.info_label.setVisible(False)
        self.info_label.setWordWrap(True)
        self.info_label.setTextFormat(getattr(Qt.TextFormat, "MarkdownText", Qt.TextFormat.AutoText))
        self.info_label.setOpenExternalLinks(True)
        self.excel_path = QLineEdit(excel_path)
        self.base_folder = QLineEdit(base_folder)
        self.zenodo_token = QLineEdit(settings.get("zenodo_token", ""))
        self.zenodo_token.setToolTip(
            "You can get token from https://zenodo.org/account/settings/applications/."
            " The token for sandbox and production are different"
        )
        self.zenodo_title = QLineEdit()
        self.zenodo_author = QLineEdit(settings.get("zenodo_author", ""))
        self.zenodo_author.setToolTip(
            "Only first author could be used from this widget. Other you need to add manually"
        )
        self.zenodo_affiliation = QLineEdit(settings.get("zenodo_affiliation", ""))
        self.zenodo_orcid = QLineEdit(settings.get("zenodo_orcid", ""))
        self.zenodo_description = QTextEdit()
        self.zenodo_description.setPlaceholderText("Put your dataset description here")
        self.zenodo_description.setSizeAdjustPolicy(QTextEdit.SizeAdjustPolicy.AdjustToContents)
        self.zenodo_sandbox = QCheckBox()
        self.zenodo_sandbox.setToolTip("Use https://sandbox.zenodo.org instead of https://zenodo.org")
        self.excel_path_btn = QPushButton("Select excel file")
        self.base_folder_btn = QPushButton("Select base folder")
        self.info_box = QTreeWidget()
        self.info_box.header().close()
        self.zenodo_collapse = QCollapsible("Zenodo export")

        self.export_btn = QPushButton("Export")
        self.export_btn.setDisabled(True)
        self.export_to_zenodo_btn = QPushButton("Export to zenodo")
        self.export_to_zenodo_btn.setDisabled(True)
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.worker = None

        self.zenodo_token.textChanged.connect(self._check_if_enable_zenodo_btn)
        self.excel_path.textChanged.connect(self._check_if_enable_zenodo_btn)
        self.base_folder.textChanged.connect(self._check_if_enable_zenodo_btn)
        self.zenodo_title.textChanged.connect(self._check_if_enable_zenodo_btn)
        self.zenodo_author.textChanged.connect(self._check_if_enable_zenodo_btn)
        self.zenodo_affiliation.textChanged.connect(self._check_if_enable_zenodo_btn)
        self.zenodo_description.textChanged.connect(self._check_if_enable_zenodo_btn)
        self.excel_path.textChanged.connect(self._excel_path_changed)
        self.base_folder.textChanged.connect(self._excel_path_changed)
        self.base_folder_btn.clicked.connect(self.select_folder)
        self.excel_path_btn.clicked.connect(self.select_excel)
        self.export_btn.clicked.connect(self._export_archive)
        self.export_to_zenodo_btn.clicked.connect(self._export_to_zenodo)

        self._setup_ui()

    def _setup_ui(self):
        layout = QGridLayout()

        layout.addWidget(self.info_label, 0, 0, 1, 3)
        layout.addWidget(QLabel("Excel file"), 1, 0)
        layout.addWidget(self.excel_path, 1, 1)
        layout.addWidget(self.excel_path_btn, 1, 2)
        layout.addWidget(QLabel("Base folder"), 2, 0)
        layout.addWidget(self.base_folder, 2, 1)
        layout.addWidget(self.base_folder_btn, 2, 2)
        layout.addWidget(self.zenodo_collapse, 3, 0, 1, 3)
        layout.addWidget(self.info_box, 4, 0, 1, 3)
        layout.addWidget(self.progress_bar, 5, 0, 1, 3)

        layout.addWidget(self.export_btn, 6, 0)
        layout.addWidget(self.export_to_zenodo_btn, 6, 2)

        self.setLayout(layout)

        zenodo_layout = QFormLayout()
        zenodo_layout.addRow("Zenodo token", self.zenodo_token)
        zenodo_layout.addRow("Dataset title", self.zenodo_title)
        zenodo_layout.addRow("Dataset author", self.zenodo_author)
        zenodo_layout.addRow("Affiliation", self.zenodo_affiliation)
        zenodo_layout.addRow("ORCID", self.zenodo_orcid)
        zenodo_layout.addRow(QLabel("Description:"))
        zenodo_layout.addRow(self.zenodo_description)
        sandbox_label = QLabel("Upload to sandbox")
        sandbox_label.setToolTip(self.zenodo_sandbox.toolTip())
        zenodo_layout.addRow(sandbox_label, self.zenodo_sandbox)

        widg = QWidget()
        widg.setLayout(zenodo_layout)
        self.zenodo_collapse.addWidget(widg)
        with suppress(KeyError):
            geometry = self.settings.get_from_profile("export_window_geometry")
            self.restoreGeometry(QByteArray.fromHex(bytes(geometry, "ascii")))

    def closeEvent(self, event) -> None:
        self.settings.set_in_profile("export_window_geometry", self.saveGeometry().toHex().data().decode("ascii"))
        super().closeEvent(event)

    def _export_archive(self):
        self.info_label.setVisible(False)
        self.export_btn.setDisabled(True)
        self.export_to_zenodo_btn.setDisabled(True)
        dlg = PSaveDialog(
            "Archive name (*.tgz *.zip *.tbz2 *.txy)",
            settings=self.settings,
            path=IO_SAVE_DIRECTORY,
            parent=self,
        )
        if dlg.exec_():
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, self.info_box.topLevelItemCount() + 1)
            self.progress_bar.setValue(0)
            export_to_archive_ = thread_worker(
                export_to_archive,
                connect={
                    "yielded": self._progress,
                    "finished": self._export_finished,
                    "errored": self._export_errored,
                    "returned": self._export_returned,
                },
                start_thread=False,
            )
            self.worker = export_to_archive_(
                excel_path=Path(self.excel_path.text()),
                base_folder=Path(self.base_folder.text()),
                target_path=Path(dlg.selectedFiles()[0]),
            )
            self.worker.start()

    def _export_finished(self):
        self.progress_bar.setVisible(False)
        self.progress_bar.setValue(0)
        self.worker = None
        self._check_if_enable_zenodo_btn()
        self._check_if_enable_export_btn()

    def _export_returned(self, result):
        self.info_label.setText(f"Export finished. {result}")
        self.info_label.setVisible(True)

    def _progress(self, value: int):
        self.progress_bar.setValue(value)

    def _export_to_zenodo(self):
        self.info_label.setVisible(False)
        self.export_btn.setDisabled(True)
        self.export_to_zenodo_btn.setDisabled(True)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, self.info_box.topLevelItemCount() + 1)
        self.progress_bar.setValue(0)

        export_to_zenodo_ = thread_worker(
            export_to_zenodo,
            connect={
                "yielded": self._progress,
                "finished": self._export_finished,
                "errored": self._export_errored,
                "returned": self._export_returned,
            },
            start_thread=False,
        )
        url = "https://zenodo.org/api/deposit/depositions"
        if self.zenodo_sandbox.isChecked():
            url = "https://sandbox.zenodo.org/api/deposit/depositions"
        self.worker = export_to_zenodo_(
            excel_path=Path(self.excel_path.text()),
            base_folder=Path(self.base_folder.text()),
            title=self.zenodo_title.text(),
            author=self.zenodo_author.text(),
            affiliation=self.zenodo_affiliation.text(),
            description=self.zenodo_description.toPlainText(),
            zenodo_token=self.zenodo_token.text(),
            orcid=self.zenodo_orcid.text(),
            zenodo_url=url,
        )
        self.settings.set("zenodo_token", self.zenodo_token.text())
        self.settings.set("zenodo_author", self.zenodo_author.text())
        self.settings.set("zenodo_affiliation", self.zenodo_affiliation.text())
        self.settings.set("zenodo_orcid", self.zenodo_orcid.text())
        self.worker.start()

    def _export_errored(self, value):
        self.info_label.setText(f"Error: {value}")
        self.info_label.setVisible(True)

    def _could_export(self):
        dir_path = Path(self.base_folder.text())
        excel_path = Path(self.excel_path.text())
        return self._all_files_exists and dir_path.is_dir() and excel_path.exists() and excel_path.is_file()

    def _check_if_enable_zenodo_btn(self):
        self.export_to_zenodo_btn.setEnabled(
            bool(
                self._could_export()
                and len(self.zenodo_token.text()) > 5
                and self.zenodo_author.text()
                and self.zenodo_affiliation.text()
                and self.zenodo_title.text()
                and self.zenodo_description.toPlainText()
            )
        )

    def _check_if_enable_export_btn(self):
        self.export_btn.setEnabled(self._could_export())

    def select_folder(self):
        dial = SelectDirectoryDialog(
            settings=self.settings,
            settings_path=[IO_BATCH_DIRECTORY, OPEN_DIRECTORY],
            default_directory=str(Path.home()),
            parent=self,
        )
        if dial.exec_():
            self.base_folder.setText(dial.selectedFiles()[0])

    def select_excel(self):
        dial = PLoadDialog(LoadPlanExcel, settings=self.settings, path=IO_SAVE_DIRECTORY)
        if dial.exec_():
            file_path = str(dial.selectedFiles()[0])
            if not os.path.splitext(file_path)[1]:
                file_path += ".xlsx"
            self.excel_path.setText(file_path)

    def _excel_path_changed(self):
        excel_path = Path(self.excel_path.text())
        if not excel_path.exists() or not excel_path.is_file():
            return
        not_icon = QIcon(os.path.join(icons_dir, "task-reject.png"))
        ok_icon = QIcon(os.path.join(icons_dir, "task-accepted.png"))

        file_and_presence_list = _extract_information_from_excel_to_export(excel_path, self.base_folder.text())
        self.info_box.clear()
        presence_all = bool(file_and_presence_list)
        if not presence_all:
            self.info_label.setText(NO_FILES)
            self.info_label.setVisible(True)
        else:
            self.info_label.setText("")
            self.info_label.setVisible(False)
            for file_path, presence in file_and_presence_list:
                widget = QTreeWidgetItem(self.info_box)
                widget.setText(0, file_path)
                if not presence:
                    widget.setIcon(0, not_icon)
                    widget.setToolTip(0, "File do not exists")
                else:
                    widget.setIcon(0, ok_icon)
                presence_all &= presence

        self._all_files_exists = presence_all
        self._check_if_enable_zenodo_btn()
        self._check_if_enable_export_btn()


def _extract_information_from_excel_to_export(
    excel_path: typing.Union[str, Path], base_folder: typing.Union[str, Path]
) -> typing.List[typing.Tuple[str, bool]]:
    """Extract information from Excel file to export"""
    file_list = []
    file_set = set()
    base_folder = Path(base_folder)

    xlsx = load_workbook(filename=excel_path, read_only=True)
    for sheet in xlsx.worksheets:
        if sheet.cell(1, 2).value != "name":
            continue
        index = 4  # offset
        while image_path := sheet.cell(index, 2).value:
            index += 1
            if image_path in file_set:
                continue
            file_set.add(image_path)
            file_list.append((image_path, (base_folder / image_path).exists()))

    return file_list


def export_to_archive(excel_path: Path, base_folder: Path, target_path: Path):
    """
    Export files to archive

    :param Path excel_path: path to Excel file
    :param Path base_folder: base folder from where paths are calculated
    :param Path target_path: path to archive
    """

    file_list = _extract_information_from_excel_to_export(excel_path, base_folder)
    if not file_list:
        raise ValueError(NO_FILES)
    if not all(presence for _, presence in file_list):
        raise ValueError(MISSING_FILES)
    ext = target_path.suffix
    if ext == ".zip":
        with zipfile.ZipFile(target_path, "w") as zip_file:
            zip_file.write(excel_path, arcname=excel_path.name)
            yield 1
            for i, (file_path, _) in enumerate(file_list, start=2):
                zip_file.write(base_folder / file_path, arcname=file_path)
                yield i

        return target_path

    mode_dict = {
        ".tgz": "w:gz",
        ".gz": "w:gz",
        ".tbz2": "w:bz2",
        ".bz2": "w:bz2",
        ".txz": "w:xz",
        ".xz": "w:xz",
        ".tar": "w:",
    }

    mode = mode_dict.get(ext)
    if mode is None:
        raise ValueError("Unknown archive type")

    with tarfile.open(target_path, mode=mode) as tar:
        tar.add(excel_path, arcname=excel_path.name)
        yield 1
        for i, (file_path, _) in enumerate(file_list, start=2):
            tar.add(base_folder / file_path, arcname=file_path)
            yield i

    return target_path


class ZenodoCreateError(ValueError):
    pass


def sleep_with_rate(response: requests.Response):
    """Sleep with rate limit"""
    if "X-RateLimit-Remaining" not in response.headers:
        return
    remaining = int(response.headers["X-RateLimit-Remaining"])
    if remaining > 0:
        return
    reset = int(response.headers["X-RateLimit-Reset"])
    sleep_time = reset - time.time()
    if sleep_time > 0:
        logging.info("Sleeping for %(sleep_time) seconds", extra={"sleep_time": sleep_time})
        time.sleep(sleep_time)


def export_to_zenodo(
    excel_path: Path,
    base_folder: Path,
    zenodo_token: str,
    title: str,
    author: str,
    affiliation: str,
    orcid: str,
    description: str,
    zenodo_url: str = "https://zenodo.org/api/deposit/depositions",
):
    """
    Export project to Zenodo

    :param excel_path: path to Excel file with output from batch processing
    :param base_folder: base folder from where paths are calculated
    :param zenodo_token: Zenodo API token
    :param title: title of the deposition
    :param author: author of the deposition
    :param affiliation: affiliation of the author
    :param orcid: ORCID of the author
    :param description: description of the deposition
    :param zenodo_url: Zenodo API URL
    :return:
    """
    file_list = _extract_information_from_excel_to_export(excel_path, base_folder)
    if not file_list:
        raise ValueError(NO_FILES)
    if not all(presence for _, presence in file_list):
        raise ValueError(MISSING_FILES)
    params = {"access_token": zenodo_token}
    headers = {"Content-Type": "application/json"}
    initial_request = requests.post(
        zenodo_url,
        params=params,
        json={},
        headers=headers,
        timeout=REQUESTS_TIMEOUT,
    )
    if initial_request.status_code != 201:
        raise ZenodoCreateError(
            "Can't create deposition. Please check your zenodo token."
            " Please remember that token for sandbox and production are different. "
            f" You could create token at "
            f"{zenodo_url.replace('api/deposit/depositions', 'account/settings/applications/')}"
            f" {initial_request.status_code} {initial_request.json()['message']}"
        )
    bucket_url = initial_request.json()["links"]["bucket"]
    deposition_id = initial_request.json()["id"]
    deposit_url = initial_request.json()["links"]["html"]

    data = {
        "metadata": {
            "title": title,
            "upload_type": "dataset",
            "creators": [{"name": author, "affiliation": affiliation}],
        }
    }

    if description:
        data["metadata"]["description"] = description
    if orcid:
        data["metadata"]["creators"][0]["orcid"] = orcid

    r = requests.put(
        f"{zenodo_url}/{deposition_id}",
        params=params,
        data=json.dumps(data),
        headers=headers,
        timeout=REQUESTS_TIMEOUT,
    )
    if r.status_code != 200:
        raise ZenodoCreateError(f"Can't update deposition metadata. Status code: {r.status_code}")

    with excel_path.open(mode="rb") as fp:
        r = requests.put(
            f"{bucket_url}/{excel_path.name}",
            data=fp,
            params=params,
            timeout=REQUESTS_TIMEOUT,
        )
        if r.status_code != 200:
            raise ZenodoCreateError(f"Can't upload excel file. Status code: {r.status_code}")
        yield 1

    for i, (filename, _) in enumerate(file_list, start=2):
        sleep_with_rate(r)
        with (base_folder / filename).open(mode="rb") as fp:
            r = requests.put(
                f"{bucket_url}/{filename}",
                data=fp,
                params=params,
                timeout=REQUESTS_TIMEOUT,
            )
            if r.status_code != 200:
                raise ZenodoCreateError(f"Can't upload file {filename}. Status code: {r.status_code}")
            yield i

    return deposit_url
