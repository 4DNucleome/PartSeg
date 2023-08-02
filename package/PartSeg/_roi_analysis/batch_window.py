import json
import logging
import multiprocessing
import os
import tarfile
import typing
import zipfile
from contextlib import suppress
from io import BytesIO
from pathlib import Path
from pickle import PicklingError  # nosec

import numpy as np
import requests
import sentry_sdk
from openpyxl.reader.excel import load_workbook
from qtpy.QtCore import QByteArray, Qt, QTimer
from qtpy.QtGui import QIcon, QStandardItem, QStandardItemModel
from qtpy.QtWidgets import (
    QCheckBox,
    QDialog,
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListView,
    QListWidgetItem,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSpinBox,
    QTabWidget,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)
from superqt.utils import thread_worker

from PartSeg import parsed_version, state_store
from PartSeg._roi_analysis.partseg_settings import PartSettings
from PartSeg._roi_analysis.prepare_plan_widget import CalculatePlaner
from PartSeg.common_backend.base_settings import IO_SAVE_DIRECTORY
from PartSeg.common_gui.custom_load_dialog import PLoadDialog, SelectDirectoryDialog
from PartSeg.common_gui.custom_save_dialog import PSaveDialog
from PartSeg.common_gui.error_report import ExceptionList, ExceptionListItem
from PartSeg.common_gui.main_window import OPEN_DIRECTORY
from PartSeg.common_gui.searchable_combo_box import SearchComboBox
from PartSeg.common_gui.select_multiple_files import IO_BATCH_DIRECTORY, AddFiles
from PartSeg.common_gui.universal_gui_part import Spacing, right_label
from PartSegCore.algorithm_describe_base import AlgorithmProperty
from PartSegCore.analysis.batch_processing.batch_backend import CalculationManager
from PartSegCore.analysis.calculation_plan import Calculation, CalculationPlan, MaskFile
from PartSegCore.io_utils import LoadPlanExcel, SaveBase
from PartSegCore.segmentation.algorithm_base import SegmentationLimitException
from PartSegCore.universal_const import Units
from PartSegData import icons_dir

__author__ = "Grzegorz Bokota"

REQUESTS_TIMEOUT = 600


class SaveExcel(SaveBase):
    @classmethod
    def get_short_name(cls):
        return "excel"

    @classmethod
    def save(
        cls,
        save_location: typing.Union[str, BytesIO, Path],
        project_info,
        parameters: dict,
        range_changed=None,
        step_changed=None,
    ):
        """empty function to satisfy interface"""

    @classmethod
    def get_name(cls) -> str:
        return "Excel (*.xlsx)"

    @classmethod
    def get_fields(cls) -> typing.List[typing.Union[AlgorithmProperty, str]]:
        return []


class ProgressView(QWidget):
    """
    :type batch_manager: CalculationManager
    """

    def __init__(self, parent, batch_manager):
        super().__init__(parent)
        self.task_count = 0
        self.calculation_manager = batch_manager
        self.whole_progress = QProgressBar(self)
        self.whole_progress.setMinimum(0)
        self.whole_progress.setMaximum(1)
        self.whole_progress.setFormat("%v of %m")
        self.whole_progress.setTextVisible(True)
        self.part_progress = QProgressBar(self)
        self.part_progress.setMinimum(0)
        self.part_progress.setMaximum(1)
        self.part_progress.setFormat("%v of %m")
        self.whole_label = QLabel("All batch progress:", self)
        self.part_label = QLabel("Single batch progress:", self)
        self.cancel_remove_btn = QPushButton("Remove task")
        self.cancel_remove_btn.setDisabled(True)
        self.logs = ExceptionList(self)
        self.logs.setToolTip("Logs")
        self.task_view = QListView()
        self.task_que = QStandardItemModel(self)
        self.task_view.setModel(self.task_que)
        self.process_num_timer = QTimer()
        self.process_num_timer.setInterval(1000)
        self.process_num_timer.setSingleShot(True)
        self.process_num_timer.timeout.connect(self.change_number_of_workers)
        self.number_of_process = QSpinBox(self)
        self.number_of_process.setRange(1, multiprocessing.cpu_count())
        self.number_of_process.setValue(1)
        self.number_of_process.setToolTip("Number of process used in batch calculation")
        self.number_of_process.valueChanged.connect(self.process_num_timer_start)
        self.progress_item_dict = {}
        self.setup_ui()
        self.preview_timer = QTimer()
        self.preview_timer.setInterval(1000)
        self.preview_timer.timeout.connect(self.update_info)
        self.task_view.selectionModel().currentChanged.connect(self.task_selection_change)
        self.cancel_remove_btn.clicked.connect(self.task_cancel_remove)

    def setup_ui(self):
        layout = QGridLayout()
        layout.addWidget(self.whole_label, 0, 0, Qt.AlignmentFlag.AlignRight)
        layout.addWidget(self.whole_progress, 0, 1, 1, 2)
        layout.addWidget(self.part_label, 1, 0, Qt.AlignmentFlag.AlignRight)
        layout.addWidget(self.part_progress, 1, 1, 1, 2)
        lab = QLabel("Number of process:")
        lab.setToolTip("Number of process used in batch calculation")
        layout.addWidget(lab, 2, 0)
        layout.addWidget(self.number_of_process, 2, 1)
        layout.addWidget(self.logs, 3, 0, 2, 3)
        layout.addWidget(self.task_view, 0, 4, 4, 1)
        layout.addWidget(self.cancel_remove_btn, 4, 4, 1, 1)
        layout.setColumnMinimumWidth(2, 10)
        layout.setColumnStretch(2, 1)
        self.setLayout(layout)

    def task_selection_change(self, new, old):
        task: CalculationProcessItem = self.task_que.item(new.row(), new.column())
        if task is None:
            self.cancel_remove_btn.setDisabled(True)
            return
        self.cancel_remove_btn.setEnabled(True)
        if task.is_finished():
            self.cancel_remove_btn.setText(f"Remove task {task.num}")
        else:
            self.cancel_remove_btn.setText(f"Cancel task {task.num}")

    def task_cancel_remove(self):
        index = self.task_view.selectionModel().currentIndex()
        task: CalculationProcessItem = typing.cast(
            CalculationProcessItem, self.task_que.item(index.row(), index.column())
        )
        if task.is_finished():
            self.calculation_manager.remove_calculation(task.calculation)
        else:
            self.calculation_manager.cancel_calculation(task.calculation)
        self.task_que.takeRow(index.row())
        print(task)

    def new_task(self):
        self.whole_progress.setMaximum(self.calculation_manager.calculation_size)
        if not self.preview_timer.isActive():
            self.update_info()
            self.preview_timer.start()

    def update_info(self):
        res = self.calculation_manager.get_results()
        for el in res.errors:
            if el[0]:
                QListWidgetItem(el[0], self.logs)
            ExceptionListItem(el[1], self.logs)
            if (
                state_store.report_errors
                and parsed_version.is_devrelease
                and not isinstance(el[1][0], SegmentationLimitException)
                and isinstance(el[1][1], tuple)
            ):
                with sentry_sdk.push_scope() as scope:
                    scope.set_tag("auto_report", "true")
                    sentry_sdk.capture_event(el[1][1][0])
        self.whole_progress.setValue(res.global_counter)
        working_search = True
        for uuid, progress in res.jobs_status.items():
            calculation = self.calculation_manager.calculation_dict[uuid]
            total = len(calculation.file_list)
            if uuid in self.progress_item_dict:
                item = self.progress_item_dict[uuid]
                item.update_count(progress)
            else:
                item = CalculationProcessItem(calculation, self.task_count, progress)
                self.task_count += 1
                self.task_que.appendRow(item)
                self.progress_item_dict[uuid] = item

            if working_search and progress != total:
                self.part_progress.setMaximum(total)
                self.part_progress.setValue(progress)
                working_search = False
        if not self.calculation_manager.has_work:
            self.part_progress.setValue(self.part_progress.maximum())
            self.preview_timer.stop()
            logging.info("Progress stop")

    def process_num_timer_start(self):
        self.process_num_timer.start()

    def update_progress(self, total_progress, part_progress):
        self.whole_progress.setValue(total_progress)
        self.part_progress.setValue(part_progress)

    def set_total_size(self, size):
        self.whole_progress.setMaximum(size)

    def set_part_size(self, size):
        self.part_progress.setMaximum(size)

    def change_number_of_workers(self):
        self.calculation_manager.set_number_of_workers(self.number_of_process.value())


class FileChoose(QWidget):
    """
    :type batch_manager: CalculationManager
    """

    def __init__(self, settings: PartSettings, batch_manager, parent=None):
        super().__init__(parent)
        self.files_to_proceed = set()
        self.settings = settings
        self.batch_manager = batch_manager
        self.files_widget = AddFiles(settings, self)
        self.progress = ProgressView(self, batch_manager)
        self.run_button = QPushButton("Process")
        self.run_button.setDisabled(True)
        self.calculation_choose = SearchComboBox()
        self.calculation_choose.addItem("<no calculation>")
        self.calculation_choose.currentTextChanged.connect(self.change_situation)
        self.result_file = QLineEdit(self)
        self.result_file.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.result_file.setReadOnly(True)
        self.choose_result = QPushButton("Save result as", self)
        self.choose_result.clicked.connect(self.choose_result_file)
        self.export_data_button = QPushButton("Export batch with data", self)
        self.export_data_button.clicked.connect(self.export_data)

        self.run_button.clicked.connect(self.prepare_calculation)
        self.files_widget.file_list_changed.connect(self.change_situation)
        self.settings.batch_plans_changed.connect(self._refresh_batch_list)

        layout = QVBoxLayout()
        layout.addWidget(self.files_widget)
        calc_layout = QHBoxLayout()
        calc_layout.addWidget(QLabel("Batch workflow:"))
        calc_layout.addWidget(self.calculation_choose)
        calc_layout.addWidget(self.result_file)
        calc_layout.addWidget(self.choose_result)
        calc_layout.addWidget(self.export_data_button)
        calc_layout.addStretch()
        calc_layout.addWidget(self.run_button)
        layout.addLayout(calc_layout)
        layout.addWidget(self.progress)
        self.setLayout(layout)

        self._refresh_batch_list()

    def export_data(self):
        dialog = ExportProjectDialog(self.result_file.text(), self.files_widget.paths_input.text(), self.settings, self)
        dialog.exec_()

    def prepare_calculation(self):
        plan = self.settings.batch_plans[str(self.calculation_choose.currentText())]
        dial = CalculationPrepare(
            self.files_widget.get_paths(), plan, str(self.result_file.text()), self.settings, self.batch_manager
        )
        if dial.exec_():
            try:
                self.batch_manager.add_calculation(dial.get_data())
                self.progress.new_task()
            except PicklingError as e:  # pragma: no cover
                if state_store.develop:
                    QMessageBox.critical(self, "Pickle error", "Please restart PartSeg.")
                else:
                    raise e

    def _refresh_batch_list(self):
        current_calc = str(self.calculation_choose.currentText())
        new_list = ["<no calculation>", *sorted(self.settings.batch_plans.keys())]
        try:
            index = new_list.index(current_calc)
        except ValueError:
            index = 0
        self.calculation_choose.clear()
        self.calculation_choose.addItems(new_list)
        self.calculation_choose.setCurrentIndex(index)

    def change_situation(self):
        if (
            str(self.calculation_choose.currentText()) == "<no calculation>"
            or len(self.files_widget.files_to_proceed) == 0
            or not str(self.result_file.text())
        ):
            self.run_button.setDisabled(True)

        else:
            self.run_button.setEnabled(True)
        if self.calculation_choose.currentText() in self.settings.batch_plans:
            plan = self.settings.batch_plans[str(self.calculation_choose.currentText())]
            self.files_widget.mask_list = plan.get_list_file_mask()
        else:
            self.files_widget.mask_list = []

    def choose_result_file(self):
        dial = PSaveDialog(SaveExcel, system_widget=False, settings=self.settings, path=IO_SAVE_DIRECTORY)
        if dial.exec_():
            file_path = str(dial.selectedFiles()[0])
            if not os.path.splitext(file_path)[1]:
                file_path += ".xlsx"
            self.result_file.setText(file_path)
            self.change_situation()


class BatchWindow(QTabWidget):
    """
    :type settings: PartSettings
    """

    def __init__(self, settings, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Batch processing")
        self.settings = settings
        self.batch_manager = CalculationManager()
        self.file_choose = FileChoose(self.settings, self.batch_manager, self)
        self.calculate_planer = CalculatePlaner(self.settings, self)
        self.addTab(self.calculate_planer, "Prepare workflow")
        self.addTab(self.file_choose, "Input files")
        self.working = False
        with suppress(KeyError):
            geometry = self.settings.get_from_profile("batch_window_geometry")
            self.restoreGeometry(QByteArray.fromHex(bytes(geometry, "ascii")))

    def focusInEvent(self, event):
        self.calculate_planer.showEvent(event)

    def is_working(self):
        return self.working

    def terminate(self):
        self.batch_manager.writer.finish()
        self.working = False

    def closeEvent(self, event):
        if self.is_working():
            ret = QMessageBox.warning(
                self,
                "Batch work",
                "Batch work is not finished. Would you like to terminate it?",
                QMessageBox.StandardButton.No | QMessageBox.StandardButton.Yes,
            )
            if ret == QMessageBox.StandardButton.Yes:
                self.terminate()
            else:
                event.ignore()
        self.settings.set_in_profile("batch_window_geometry", self.saveGeometry().toHex().data().decode("ascii"))
        super().closeEvent(event)


class CalculationPrepare(QDialog):
    """
    :type mask_path_list: list[QLineEdit]
    :type mask_mapper_list: list[MaskMapper]
    """

    def __init__(
        self,
        file_list: typing.List[os.PathLike],
        calculation_plan: CalculationPlan,
        measurement_file_path: os.PathLike,
        settings: PartSettings,
        batch_manager: CalculationManager,
        parent: typing.Optional[QWidget] = None,
    ):
        """
        :param file_list: list of files to proceed
        :param calculation_plan: calculation plan for this run
        :param measurement_file_path: path to measurement result file
        :param settings: settings object
        :type batch_manager: CalculationManager
        """
        super().__init__(parent=parent)
        self.setWindowTitle("Calculation start")
        self.file_list = file_list
        self.calculation_plan = calculation_plan
        self.measurement_file_path = measurement_file_path
        self.settings = settings
        self.batch_manager = batch_manager
        self.info_label = QLabel(
            f"Information, <i><font color='{self._warning_color()}'>warnings</font></i>, "
            "<b><font color='red'>errors</font><b>"
        )
        self.voxel_size = Spacing("Voxel size", settings.image.spacing, settings.get("units_value", Units.nm))
        if len(file_list) == 1:
            all_prefix = os.path.dirname(file_list[0])
        else:
            all_prefix = os.path.commonpath(file_list)
        self.all_file_prefix = all_prefix
        self.base_prefix = QLineEdit(all_prefix, self)
        self.base_prefix.setReadOnly(True)
        self.result_prefix = QLineEdit(all_prefix, self)
        self.result_prefix.setReadOnly(True)
        self.base_prefix_btn = QPushButton("Choose data prefix")
        self.base_prefix_btn.clicked.connect(self.choose_data_prefix)
        self.result_prefix_btn = QPushButton("Choose save prefix")
        self.result_prefix_btn.clicked.connect(self.choose_result_prefix)
        self.sheet_name = QLineEdit("Sheet1")
        self.sheet_name.textChanged.connect(self.verify_data)
        self.measurement_file_path_view = QLineEdit(str(measurement_file_path))
        self.measurement_file_path_view.setReadOnly(True)

        self.overwrite_voxel_size_check = QCheckBox("Overwrite voxel size")
        self.overwrite_voxel_size_check.stateChanged.connect(self._overwrite_voxel_size_check_changed)

        self.mask_path_list = []
        self.mask_mapper_list = self.calculation_plan.get_list_file_mask()
        mask_file_list = [(i, el) for i, el in enumerate(self.mask_mapper_list) if isinstance(el, MaskFile)]

        self.state_list = np.zeros((len(self.file_list), len(self.mask_mapper_list)), dtype=np.uint8)

        self.file_list_widget = QTreeWidget()
        self.file_list_widget.header().close()

        self.execute_btn = QPushButton("Execute")
        self.execute_btn.clicked.connect(self.accept)
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.close)

        self.setup_ui(mask_file_list)
        self.verify_data()

    def setup_ui(self, mask_file_list):
        mask_path_layout = QGridLayout()
        for i, (pos, mask_file) in enumerate(mask_file_list):
            if not mask_file.name:
                mask_path_layout.addWidget(right_label(f"Path to file {i + 1} with mask mapping"))
            else:
                mask_path_layout.addWidget(
                    right_label(f"Path to file {i + 1} with mask mapping for name: {mask_file.name}")
                )
            mask_path = QLineEdit(self)
            mask_path.setReadOnly(True)
            self.mask_path_list.append(mask_path)
            set_path = QPushButton("Choose file", self)
            set_path.clicked.connect(self.set_mapping_mask(i, pos))
            mask_path_layout.addWidget(mask_path, i, 1)
            mask_path_layout.addWidget(set_path, i, 2)

        layout = QGridLayout()
        layout.addWidget(self.info_label, 0, 0, 1, 5)
        layout.addWidget(self.voxel_size, 1, 0, 1, 5)
        layout.addWidget(self.overwrite_voxel_size_check, 2, 0, 1, 5)
        layout.addWidget(right_label("Measurement sheet name:"), 4, 3)
        layout.addWidget(self.sheet_name, 4, 4)
        layout.addWidget(right_label("Measurement file path:"), 3, 3)
        layout.addWidget(self.measurement_file_path_view, 3, 4)

        layout.addWidget(right_label("Data prefix:"), 3, 0)
        layout.addWidget(self.base_prefix, 3, 1)
        layout.addWidget(self.base_prefix_btn, 3, 2)
        layout.addWidget(right_label("Save prefix:"), 4, 0)
        layout.addWidget(self.result_prefix, 4, 1)
        layout.addWidget(self.result_prefix_btn, 4, 2)
        layout.addLayout(mask_path_layout, 5, 0, 1, 0)

        layout.addWidget(self.file_list_widget, 5, 0, 3, 6)
        btn_layout = QHBoxLayout()
        btn_layout.addWidget(self.execute_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(self.cancel_btn)
        layout.addLayout(btn_layout, 8, 0, 1, 0)
        self.setLayout(layout)

    def _warning_color(self):
        return "yellow" if self.settings.theme_name == "dark" else "blue"

    def _overwrite_voxel_size_check_changed(self):
        self.verify_data()
        if self.overwrite_voxel_size_check.isChecked():
            text = self.info_label.text()
            text += "<br><strong>Overwrite voxel size is checked. File metadata will be ignored</strong>"
            self.info_label.setText(text)

    def choose_data_prefix(self):
        dial = QFileDialog()
        dial.setAcceptMode(QFileDialog.AcceptMode.AcceptOpen)
        dial.setFileMode(QFileDialog.FileMode.Directory)
        dial.setDirectory(self.base_prefix.text())
        dial.setHistory(dial.history() + self.settings.get_path_history())
        if dial.exec_():
            dir_path = str(dial.selectedFiles()[0])
            self.base_prefix.setText(dir_path)

    def choose_result_prefix(self):
        dial = QFileDialog()
        dial.setOption(QFileDialog.Option.DontUseNativeDialog, True)
        dial.setAcceptMode(QFileDialog.AcceptMode.AcceptOpen)
        dial.setFileMode(QFileDialog.FileMode.Directory)
        dial.setDirectory(self.result_prefix.text())
        dial.setHistory(dial.history() + self.settings.get_path_history())
        if dial.exec_():
            dir_path = str(dial.selectedFiles()[0])
            self.result_prefix.setText(dir_path)

    def set_mapping_mask(self, i, pos):
        def mapping_dialog():
            dial = QFileDialog(self, "Select file")
            dial.setHistory(dial.history() + self.settings.get_path_history())
            base_path = str(self.base_prefix.text()).strip()
            if base_path:
                dial.setDirectory(base_path)
            dial.setFileMode(QFileDialog.FileMode.ExistingFile)
            if dial.exec_():
                path = str(dial.selectedFiles())
                self.mask_path_list[i].setText(path)
                file_mapper: MaskFile = self.mask_mapper_list[pos]
                file_mapper.set_map_path(path)

        return mapping_dialog

    def get_data(self):
        res = {
            "file_list": self.file_list,
            "base_prefix": str(self.base_prefix.text()),
            "result_prefix": str(self.result_prefix.text()),
            "measurement_file_path": str(self.measurement_file_path_view.text()),
            "sheet_name": str(self.sheet_name.text()),
            "calculation_plan": self.calculation_plan,
            "voxel_size": self.voxel_size.get_values(),
            "overwrite_voxel_size": self.overwrite_voxel_size_check.isChecked(),
        }
        return Calculation(**res)

    def verify_data(self):
        self.execute_btn.setEnabled(True)
        warning_color = self._warning_color()
        text = (
            f"information, <i><font color='{warning_color}'>warnings</font></i>,"
            f" <b><font color='red'>errors</font></b><br>"
            "The voxel size is for file in which metadata do not contains this information<br>"
        )
        if not self.batch_manager.is_valid_sheet_name(
            str(self.measurement_file_path_view.text()), str(self.sheet_name.text())
        ):
            text += f"<i><font color='{warning_color}'>Sheet name already in use</i></font><br>"
            self.execute_btn.setDisabled(True)
        if self.state_list.size > 0:
            val = np.unique(self.state_list)
            if 1 in val:
                self.execute_btn.setDisabled(True)
                text += f"<i><font color='{warning_color}'>Some mask map file are not set</font></i><br>"
            if 2 in val:
                self.execute_btn.setDisabled(True)
                text += "<b><font color='red'>Some mask do not exists</font><b><br>"

        if not all(os.path.exists(f) for f in self.file_list):
            self.execute_btn.setDisabled(True)
            text += "<b><font color='red'>Some files do not exists</font><b><br>"

        text = text[:-4]
        self.info_label.setText(text)

    def _check_start_conditions(self):
        for file_num, file_path in enumerate(self.file_list):
            for mask_num, mask_mapper in enumerate(self.mask_mapper_list):
                if mask_mapper.is_ready():
                    mask_path = mask_mapper.get_mask_path(file_path)
                    if os.path.exists(mask_path):
                        self.state_list[file_num, mask_num] = 0
                    else:
                        self.state_list[file_num, mask_num] = 2
                else:
                    self.state_list[file_num, mask_num] = 1
        self.verify_data()

    def showEvent(self, event):
        super().showEvent(event)
        self._check_start_conditions()

        icon_dkt = {
            0: QIcon(os.path.join(icons_dir, "task-accepted.png")),
            1: QIcon(os.path.join(icons_dir, "task-reject.png")),
            2: QIcon(os.path.join(icons_dir, "task-attempt.png")),
        }

        text_dkt = {
            0: "Mask {} ok",
            1: "Mask {} unknown",
            2: "Mask {} file does not exists",
        }

        warn_state = np.amax(self.state_list, axis=1, initial=0)
        for file_num, file_path in enumerate(self.file_list):
            widget = QTreeWidgetItem(self.file_list_widget)
            widget.setText(0, os.path.relpath(file_path, self.all_file_prefix))
            if not os.path.exists(file_path):
                widget.setIcon(0, icon_dkt[0])
                widget.setToolTip(0, "File do not exists")
                continue
            for mask_num, mask_mapper in enumerate(self.mask_mapper_list):
                sub_widget = QTreeWidgetItem(widget)
                sub_widget.setText(0, text_dkt[self.state_list[file_num, mask_num]].format(mask_mapper.name))
                sub_widget.setIcon(0, icon_dkt[self.state_list[file_num, mask_num]])

            widget.setIcon(0, icon_dkt[warn_state[file_num]])


class ExportProjectDialog(QDialog):
    """Export data for zenodo"""

    def __init__(
        self, excel_path: str, base_folder: str, settings: PartSettings, parent: typing.Optional[QWidget] = None
    ) -> None:
        super().__init__(parent=parent)
        self.setWindowTitle("Export batch with data")
        self.settings = settings
        self._all_files_exists = False
        self.info_label = QLabel()
        self.info_label.setVisible(False)
        self.excel_path = QLineEdit(excel_path)
        self.base_folder = QLineEdit(base_folder)
        self.zenodo_token = QLineEdit(settings.get("zenodo_token", ""))
        self.excel_path_btn = QPushButton("Select excel file")
        self.base_folder_btn = QPushButton("Select base folder")
        self.info_box = QTreeWidget()
        self.info_box.header().close()

        self.export_btn = QPushButton("Export")
        self.export_btn.setDisabled(True)
        self.export_to_zenodo_btn = QPushButton("Export to zenodo")
        self.export_to_zenodo_btn.setDisabled(True)
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.worker = None

        layout = QGridLayout()

        layout.addWidget(self.info_label, 0, 0, 1, 3)
        layout.addWidget(QLabel("Excel file"), 1, 0)
        layout.addWidget(self.excel_path, 1, 1)
        layout.addWidget(self.excel_path_btn, 1, 2)
        layout.addWidget(QLabel("Base folder"), 2, 0)
        layout.addWidget(self.base_folder, 2, 1)
        layout.addWidget(self.base_folder_btn, 2, 2)
        layout.addWidget(QLabel("Zenodo token"), 3, 0)
        layout.addWidget(self.zenodo_token, 3, 1, 1, 2)
        layout.addWidget(self.info_box, 4, 0, 1, 3)
        layout.addWidget(self.progress_bar, 5, 0, 1, 3)

        layout.addWidget(self.export_btn, 6, 0)
        layout.addWidget(self.export_to_zenodo_btn, 6, 2)

        self.setLayout(layout)

        self.zenodo_token.textChanged.connect(self._zenodo_token_refresh)
        self.excel_path.textChanged.connect(self._zenodo_token_refresh)
        self.base_folder.textChanged.connect(self._zenodo_token_refresh)
        self.excel_path.textChanged.connect(self._excel_path_changed)
        self.base_folder.textChanged.connect(self._excel_path_changed)
        self.base_folder_btn.clicked.connect(self.select_folder)
        self.excel_path_btn.clicked.connect(self.select_excel)
        self.export_btn.clicked.connect(self._export_archive)
        self.export_to_zenodo_btn.clicked.connect(self._export_to_zenodo)

    def _export_archive(self):
        dlg = PSaveDialog(
            "Archive name (*.tgz *.zip *.tbz2 *.txy)",
            settings=self.settings,
            path=IO_SAVE_DIRECTORY,
            parent=self,
        )
        if dlg.exec_():
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, self.info_box.topLevelItemCount() + 1)
            self.progress_bar.setValue(0)
            export_to_archive_ = thread_worker(export_to_archive)
            self.worker = export_to_archive_(
                excel_path=Path(self.excel_path.text()),
                base_folder=Path(self.base_folder.text()),
                target_path=Path(dlg.selectedFiles()[0]),
            )
            self.worker.yielded.connect(self._progress)
            self.worker.finished.connect(self._export_finished)
            self.worker.errored.connect(self._export_errored)
            self.worker.start()

    def _export_finished(self):
        self.progress_bar.setVisible(False)
        self.progress_bar.setValue(0)
        self.worker = None

    def _progress(self, value: int):
        self.progress_bar.setValue(value)

    def _export_to_zenodo(self):
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, self.info_box.topLevelItemCount() + 1)
        self.progress_bar.setValue(0)

        export_to_zenodo_ = thread_worker(export_to_zenodo)
        self.worker = export_to_zenodo_(
            Path(self.excel_path.text()), Path(self.base_folder.text()), zenodo_token=self.zenodo_token.text()
        )
        self.worker.yielded.connect(self._progress)
        self.worker.finished.connect(self._export_finished)
        self.worker.returned.connect(self._export_finished)
        self.worker.errored.connect(self._export_errored)
        self.worker.start()

    def _export_errored(self, value):
        self.info_label.setText(f"Error: {value}")
        self.info_label.setVisible(True)

    def _zenodo_export_finished(self, value):
        deposition_id, deposit_url = value
        self.info_label.setText(f"Deposition id: {deposition_id}\nDeposit url: {deposit_url}")
        self.info_label.setVisible(True)

    def _could_export(self):
        dir_path = Path(self.base_folder.text())
        excel_path = Path(self.excel_path.text())
        return self._all_files_exists and dir_path.is_dir() and excel_path.exists() and excel_path.is_file()

    def _zenodo_token_refresh(self):
        self.export_to_zenodo_btn.setEnabled(self._could_export() and len(self.zenodo_token.text()) > 5)

    def _export_btn_refresh(self):
        self.export_btn.setEnabled(self._could_export())

    def select_folder(self):
        dial = SelectDirectoryDialog(
            settings=self.settings,
            path=[IO_BATCH_DIRECTORY, OPEN_DIRECTORY],
            default_directory=str(Path.home()),
            parent=self,
        )
        if dial.exec_():
            self.base_folder.setText(dial.selectedFiles()[0])

    def select_excel(self):
        dial = PLoadDialog(LoadPlanExcel, settings=self.settings, path=IO_SAVE_DIRECTORY)
        if dial.exec_():
            file_path = str(dial.selectedFiles()[0])
            if not os.path.splitext(file_path)[1]:
                file_path += ".xlsx"
            self.excel_path.setText(file_path)

    def _excel_path_changed(self):
        excel_path = Path(self.excel_path.text())
        if not excel_path.exists() or not excel_path.is_file():
            return
        not_icon = QIcon(os.path.join(icons_dir, "task-reject.png"))
        ok_icon = QIcon(os.path.join(icons_dir, "task-accepted.png"))

        file_and_presence_list = _extract_information_from_excel_to_export(excel_path, self.base_folder.text())
        self.info_box.clear()
        presence_all = bool(file_and_presence_list)
        if not presence_all:
            self.info_label.setText("No files to export")
            self.info_label.setVisible(True)
        else:
            self.info_label.setText("")
            self.info_label.setVisible(False)
            for file_path, presence in file_and_presence_list:
                widget = QTreeWidgetItem(self.info_box)
                widget.setText(0, file_path)
                if not presence:
                    widget.setIcon(0, not_icon)
                    widget.setToolTip(0, "File do not exists")
                else:
                    widget.setIcon(0, ok_icon)
                presence_all &= presence

        self._all_files_exists = presence_all
        self._zenodo_token_refresh()
        self._export_btn_refresh()


def _extract_information_from_excel_to_export(
    excel_path: typing.Union[str, Path], base_folder: typing.Union[str, Path]
) -> typing.List[typing.Tuple[str, bool]]:
    """Extract information from Excel file to export"""
    file_list = []
    file_set = set()
    base_folder = Path(base_folder)

    xlsx = load_workbook(filename=excel_path, read_only=True)
    for sheet in xlsx.worksheets:
        if sheet.cell(1, 2).value != "name":
            continue
        index = 4  # offset
        while image_path := sheet.cell(index, 2).value:
            index += 1
            if image_path in file_set:
                continue
            file_set.add(image_path)
            file_list.append((image_path, (base_folder / image_path).exists()))

    return file_list


def export_to_archive(excel_path: Path, base_folder: Path, target_path: Path):
    """
    Export files to archive

    :param Path excel_path: path to excel file
    :param Path base_folder: base folder from where paths are calculated
    :param Path target_path: path to archive
    """

    file_list = _extract_information_from_excel_to_export(excel_path, base_folder)
    if not file_list:
        raise ValueError("No files to export")
    if not all(presence for _, presence in file_list):
        raise ValueError("Some files do not exists")
    ext = target_path.suffix
    if ext == ".zip":
        with zipfile.ZipFile(target_path, "w") as zip_file:
            zip_file.write(excel_path, arcname=excel_path.name)
            yield 1
            for i, (file_path, _) in enumerate(file_list, start=2):
                zip_file.write(base_folder / file_path, arcname=file_path)
                yield i

        return

    mode_dict = {
        ".tgz": "w:gz",
        ".gz": "w:gz",
        ".tbz2": "w:bz2",
        ".bz2": "w:bz2",
        ".txz": "w:xz",
        ".xz": "w:xz",
        ".tar": "w:",
    }

    mode = mode_dict.get(ext)
    if mode is None:
        raise ValueError("Unknown archive type")

    with tarfile.open(target_path, mode=mode) as tar:
        tar.add(excel_path, arcname=excel_path.name)
        yield 1
        for i, (file_path, _) in enumerate(file_list, start=2):
            tar.add(base_folder / file_path, arcname=file_path)
            yield i


def export_to_zenodo(
    excel_path: Path,
    base_folder: Path,
    zenodo_token: str,
    zenodo_url: str = "https://sandbox.zenodo.org/api/deposit/depositions",
    # 'https://zenodo.org/api/deposit/depositions'
):
    """
    Export project to Zenodo

    :param excel_path:
    :param base_folder:
    :param zenodo_token:
    :param zenodo_url: Zenodo API URL
    :return:
    """
    file_list = _extract_information_from_excel_to_export(excel_path, base_folder)
    if not file_list:
        raise ValueError("No files to export")
    if not all(presence for _, presence in file_list):
        raise ValueError("Some files do not exists")
    params = {"access_token": zenodo_token}
    headers = {"Content-Type": "application/json"}
    initial_request = requests.post(
        "https://sandbox.zenodo.org/api/deposit/depositions",
        params=params,
        json={},
        headers=headers,
        timeout=REQUESTS_TIMEOUT,
    )
    if initial_request.status_code != 201:
        raise ValueError(f"Can't create deposition {initial_request.status_code} {initial_request.json()['message']}")
    bucket_url = initial_request.json()["links"]["bucket"]
    deposition_id = initial_request.json()["id"]
    deposit_url = initial_request.json()["links"]["html"]

    data = {
        "metadata": {
            "title": "My first PartSeg upload",
            "upload_type": "dataset",
            "description": "Upload data from PartSeg",
            "creators": [{"name": "Grzegorz Bokota", "affiliation": "PartSeg"}],
        }
    }
    requests.put(
        f"{zenodo_url}/{deposition_id}",
        params=params,
        data=json.dumps(data),
        headers=headers,
        timeout=REQUESTS_TIMEOUT,
    )
    # r.status_code

    with excel_path.open(mode="rb") as fp:
        requests.put(
            f"{bucket_url}/{excel_path.name}",
            data=fp,
            params=params,
            timeout=REQUESTS_TIMEOUT,
        )
        yield 1

    for i, (filename, _) in enumerate(file_list, start=2):
        with (base_folder / filename).open(mode="rb") as fp:
            requests.put(
                f"{bucket_url}/{filename}",
                data=fp,
                params=params,
                timeout=REQUESTS_TIMEOUT,
            )
            yield i

    return deposition_id, deposit_url


class CalculationProcessItem(QStandardItem):
    def __init__(self, calculation: Calculation, num: int, count, *args, **kwargs):
        text = f"Task {num} ({count}/{len(calculation.file_list)})"
        super().__init__(text, *args, **kwargs)
        self.calculation = calculation
        self.num = num
        self.count = count
        self.setToolTip(str(calculation.calculation_plan))
        self.setEditable(False)

    def update_count(self, count):
        self.count = count
        self.setText(f"Task {self.num} ({count}/{len(self.calculation.file_list)})")

    def is_finished(self) -> bool:
        return self.count == len(self.calculation.file_list)
