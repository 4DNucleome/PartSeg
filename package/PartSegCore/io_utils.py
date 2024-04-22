import io
import json
import os
import re
import typing
from abc import ABC, abstractmethod
from datetime import datetime
from enum import Enum
from io import BufferedIOBase, BytesIO, IOBase, RawIOBase, StringIO, TextIOBase
from pathlib import Path
from tarfile import TarFile, TarInfo

import imageio
import numpy as np
import pandas as pd
import tifffile
from openpyxl import load_workbook

from PartSegCore.algorithm_describe_base import AlgorithmDescribeBase, AlgorithmProperty
from PartSegCore.json_hooks import partseg_object_hook
from PartSegCore.project_info import ProjectInfoBase
from PartSegCore.utils import EventedDict, ProfileDict, check_loaded_dict, iterate_names
from PartSegImage import ImageWriter
from PartSegImage.image import minimal_dtype


class SegmentationType(Enum):
    analysis = 1
    mask = 2


class WrongFileTypeException(Exception):
    pass


class NotSupportedImage(Exception):
    pass


def check_segmentation_type(tar_file: TarFile) -> SegmentationType:
    names = [x.name for x in tar_file.getmembers()]
    if "algorithm.json" in names:
        return SegmentationType.analysis
    if "metadata.json" in names:
        return SegmentationType.mask
    raise WrongFileTypeException  # pragma: no cover


def get_tarinfo(name, buffer: typing.Union[BytesIO, StringIO]):
    tar_info = TarInfo(name=name)
    buffer.seek(0)
    if isinstance(buffer, BytesIO):
        tar_info.size = len(buffer.getbuffer())
    else:
        tar_info.size = len(buffer.getvalue())
    tar_info.mtime = datetime.now().timestamp()
    return tar_info


class _IOBase(AlgorithmDescribeBase, ABC):
    @classmethod
    def get_name_with_suffix(cls):
        return cls.get_name()

    @classmethod
    def get_extensions(cls) -> typing.List[str]:
        match = re.match(r".*\((.*)\)", cls.get_name())
        if match is None:
            raise ValueError(f"No extensions found in {cls.get_name()}")
        extensions = match[1].split(" ")
        if not all(x.startswith("*.") for x in extensions):
            raise ValueError(f"Error with parsing extensions in {cls.get_name()}")
        return [x[1:] for x in extensions]


class SaveBase(_IOBase, ABC):
    need_functions: typing.ClassVar[typing.List[str]] = [
        "save",
        "get_short_name",
        "get_name_with_suffix",
        "get_default_extension",
        "need_segmentation",
        "need_mask",
    ]

    @classmethod
    @abstractmethod
    def get_short_name(cls):
        raise NotImplementedError

    @classmethod
    @abstractmethod
    def save(
        cls,
        save_location: typing.Union[str, BytesIO, Path],
        project_info,
        parameters: dict,
        range_changed=None,
        step_changed=None,
    ):
        """

        :param save_location: location to save, can be buffer
        :param project_info: all project data
        :param parameters: additional parameters for saving method
        :param range_changed: report function for inform about steps num
        :param step_changed: report function for progress
        """
        raise NotImplementedError

    @classmethod
    def need_segmentation(cls):
        return True

    @classmethod
    def need_mask(cls):
        return False

    @classmethod
    def get_default_extension(cls):
        match = re.search(r"\(\*(\.\w+)", cls.get_name_with_suffix())
        return match[1] if match else ""


class LoadBase(_IOBase, ABC):
    need_functions: typing.ClassVar[typing.List[str]] = [
        "load",
        "get_short_name",
        "get_name_with_suffix",
        "number_of_files",
        "get_next_file",
        "partial",
    ]

    @classmethod
    @abstractmethod
    def get_short_name(cls):
        raise NotImplementedError

    @classmethod
    @abstractmethod
    def load(
        cls,
        load_locations: typing.List[typing.Union[str, BytesIO, Path]],
        range_changed: typing.Optional[typing.Callable[[int, int], typing.Any]] = None,
        step_changed: typing.Optional[typing.Callable[[int], typing.Any]] = None,
        metadata: typing.Optional[dict] = None,
    ) -> typing.Union[ProjectInfoBase, typing.List[ProjectInfoBase]]:
        """
        Function for load data

        :param load_locations: list of files to load
        :param range_changed: callback function for inform about number of steps to be done
        :param step_changed:  callback function for report that single step has been done
        :param metadata: additional information needed by function. Like default spacing for load image
        :return: Project info or list of project info
        """
        raise NotImplementedError

    @classmethod
    def get_fields(cls):
        return []

    @classmethod
    def number_of_files(cls):
        """Number of files required for load method"""
        return 1

    @classmethod
    def get_next_file(cls, file_paths: typing.List[str]):
        return file_paths[0]

    @classmethod
    def partial(cls):
        """Inform that this class load complete data"""
        return False


def load_metadata_base(data: typing.Union[str, Path, typing.TextIO]):
    try:
        if isinstance(data, io.TextIOBase):
            decoded_data = json.load(data, object_hook=partseg_object_hook)
        elif os.path.exists(data):
            with open(data, encoding="utf-8") as ff:
                decoded_data = json.load(ff, object_hook=partseg_object_hook)
        else:
            decoded_data = json.loads(data, object_hook=partseg_object_hook)
    except ValueError as e:  # pragma: no cover
        try:
            decoded_data = json.loads(str(data), object_hook=partseg_object_hook)
        except Exception:
            raise e  # noqa: B904

    return decoded_data


def load_metadata_part(data: typing.Union[str, Path]) -> typing.Tuple[typing.Any, typing.List[typing.Tuple[str, dict]]]:
    """
    Load serialized data. Get valid entries.

    :param data: path to file or string to be decoded.
    :return:
    """
    # TODO extract to function
    data = load_metadata_base(data)
    bad_key = []
    if isinstance(data, typing.MutableMapping) and "__error__" in data:
        bad_key.append(data)
        data = {}
    if isinstance(data, typing.MutableMapping) and not check_loaded_dict(data):
        bad_key.extend((k, data.pop(k)) for k, v in list(data.items()) if not check_loaded_dict(v))
    elif isinstance(data, ProfileDict) and not data.verify_data():
        bad_key = data.pop_errors()
    return data, bad_key


load_matadata_part = load_metadata_part
# backward compatibility


def find_problematic_entries(data: typing.Any) -> typing.List[typing.MutableMapping]:
    """
    Find top nodes with ``"__error__"`` key. If node found
    then its children is not checked.

    :param data: data to be checked
    :return:  top level entries with "__error__" key
    """
    if not isinstance(data, typing.MutableMapping):
        return []
    if "__error__" in data:
        return [data]
    res = []
    for v in data.values():
        res.extend(find_problematic_entries(v))
    return res


def find_problematic_leafs(data: typing.Any) -> typing.List[typing.MutableMapping]:
    """
    Find bottom nodes with ``"__error__"`` key. If any
    children has ``"__error__"`` then such node is not returned.

    :param data: data to be checked.
    :return: bottom level entries with "__error__" key
    """
    if not isinstance(data, typing.MutableMapping):
        return []
    if "__error__" not in data and (not isinstance(data, EventedDict) or len(data) == 0):
        return []
    res = []
    data_to_check = data
    if "__class__" in data and "__values__" in data:
        data_to_check = data["__values__"]
    for data_ in data_to_check.values():
        res.extend(find_problematic_leafs(data_))
    return res or [data]


def proxy_callback(
    range_changed: typing.Callable[[int, int], typing.Any],
    step_changed: typing.Callable[[int], typing.Any],
    text: str,
    val,
):
    if text == "max" and range_changed is not None:
        range_changed(0, val)
    if text == "step" and step_changed is not None:
        step_changed(val)


def open_tar_file(
    file_data: typing.Union[str, Path, TarFile, TextIOBase, BufferedIOBase, RawIOBase, IOBase], mode="r"
) -> typing.Tuple[TarFile, str]:
    """Create tar file from path or buffer. If passed :py:class:`TarFile` then return it."""
    if isinstance(file_data, TarFile):
        tar_file = file_data
        file_path = ""
    elif isinstance(file_data, (str, Path)):
        tar_file = TarFile.open(file_data, mode)
        file_path = str(file_data)
    elif isinstance(file_data, (TextIOBase, BufferedIOBase, RawIOBase, IOBase)):
        tar_file = TarFile.open(fileobj=file_data)
        file_path = ""
    else:
        raise ValueError(f"wrong type of file_data argument: {type(file_data)}")
    return tar_file, file_path


class SaveMaskAsTiff(SaveBase):
    @classmethod
    def get_name(cls):
        return "Mask (*.tiff *.tif)"

    @classmethod
    def get_short_name(cls):
        return "mask_tiff"

    @classmethod
    def get_fields(cls):
        return []

    @classmethod
    def need_mask(cls):
        return True

    @classmethod
    def save(
        cls,
        save_location: typing.Union[str, BytesIO, Path],
        project_info,
        parameters: typing.Optional[dict] = None,
        range_changed=None,
        step_changed=None,
    ):
        if project_info.image.mask is None and project_info.mask is not None:
            ImageWriter.save_mask(project_info.image.substitute(mask=project_info.mask), save_location)
        else:
            ImageWriter.save_mask(project_info.image, save_location)


def tar_to_buff(tar_file, member_name) -> BytesIO:
    tar_value = tar_file.extractfile(tar_file.getmember(member_name))
    buffer = BytesIO()
    buffer.write(tar_value.read())
    buffer.seek(0)
    return buffer


class SaveScreenshot(SaveBase):
    @classmethod
    def get_short_name(cls):
        return "screenshot"

    @classmethod
    def save(
        cls,
        save_location: typing.Union[str, BytesIO, Path],
        project_info,
        parameters: typing.Optional[dict] = None,
        range_changed=None,
        step_changed=None,
    ):
        imageio.imsave(save_location, project_info)

    @classmethod
    def get_name(cls) -> str:
        return "Screenshot (*.png *.jpg *.jpeg)"

    @classmethod
    def get_fields(cls) -> typing.List[typing.Union[AlgorithmProperty, str]]:
        return []


class SaveROIAsTIFF(SaveBase):
    @classmethod
    def get_name(cls):
        return "ROI as tiff (*.tiff *.tif)"

    @classmethod
    def get_short_name(cls):
        return "roi_tiff"

    @classmethod
    def get_fields(cls):
        return []

    @classmethod
    def save(
        cls,
        save_location: typing.Union[str, BytesIO, Path],
        project_info,
        parameters: dict,
        range_changed=None,
        step_changed=None,
    ):
        roi = project_info.roi_info.roi
        roi_max = max(project_info.roi_info.bound_info)
        roi = roi.astype(minimal_dtype(roi_max))
        tifffile.imwrite(save_location, roi)


class SaveROIAsNumpy(SaveBase):
    @classmethod
    def get_name(cls):
        return "ROI as numpy (*.npy)"

    @classmethod
    def get_short_name(cls):
        return "ROI_numpy"

    @classmethod
    def get_fields(cls):
        return []

    @classmethod
    def save(
        cls,
        save_location: typing.Union[str, BytesIO, Path],
        project_info,
        parameters: typing.Optional[dict] = None,
        range_changed=None,
        step_changed=None,
    ):
        roi = project_info.roi_info.roi
        roi_max = max(project_info.roi_info.bound_info)
        roi = roi.astype(minimal_dtype(roi_max))
        np.save(save_location, roi)


class PointsInfo(typing.NamedTuple):
    file_path: str
    points: np.ndarray


class LoadPoints(LoadBase):
    @classmethod
    def get_short_name(cls):
        return "point_csv"

    @classmethod
    def load(
        cls,
        load_locations: typing.List[typing.Union[str, BytesIO, Path]],
        range_changed: typing.Optional[typing.Callable[[int, int], typing.Any]] = None,
        step_changed: typing.Optional[typing.Callable[[int], typing.Any]] = None,
        metadata: typing.Optional[dict] = None,
    ) -> PointsInfo:
        df = pd.read_csv(load_locations[0], delimiter=",", index_col=0)
        return PointsInfo(load_locations[0], df.to_numpy())

    @classmethod
    def get_name(cls) -> str:
        return "Points (*.csv)"

    @classmethod
    def get_fields(cls) -> typing.List[typing.Union[AlgorithmProperty, str]]:
        return ["text"]

    @classmethod
    def partial(cls):
        return True


class LoadPlanJson(LoadBase):
    @classmethod
    def get_short_name(cls):
        return "plan_json"

    @classmethod
    def load(
        cls,
        load_locations: typing.List[typing.Union[str, BytesIO, Path]],
        range_changed: typing.Optional[typing.Callable[[int, int], typing.Any]] = None,
        step_changed: typing.Optional[typing.Callable[[int], typing.Any]] = None,
        metadata: typing.Optional[dict] = None,
    ):
        from PartSegCore.analysis.calculation_plan import CalculationPlan

        res, err = load_metadata_part(load_locations[0])
        res_dkt = {}
        err_li = []
        for key, value in res.items():
            if isinstance(value, CalculationPlan) and value.is_bad():
                err_li.append(f"Problem with load {value.name} because of {value.get_error_source()}")
            else:
                res_dkt[key] = value
        return res_dkt, err + err_li

    @classmethod
    def get_name(cls) -> str:
        return "Calculation plans (*.json)"


class LoadPlanExcel(LoadBase):
    @classmethod
    def get_short_name(cls):
        return "plan_excel"

    @classmethod
    def load(
        cls,
        load_locations: typing.List[typing.Union[str, BytesIO, Path]],
        range_changed: typing.Optional[typing.Callable[[int, int], typing.Any]] = None,
        step_changed: typing.Optional[typing.Callable[[int], typing.Any]] = None,
        metadata: typing.Optional[dict] = None,
    ):
        data_list, error_list = [], []

        xlsx = load_workbook(filename=load_locations[0], read_only=True)
        try:
            for sheet_name in xlsx.sheetnames:
                if sheet_name.startswith("info"):
                    data = ""
                    index = 2  # skip header
                    while xlsx[sheet_name].cell(row=index, column=2).value:
                        data += xlsx[sheet_name].cell(row=index, column=2).value
                        index += 1

                    try:
                        data, err = load_metadata_part(data)
                        data_list.append(data)
                        error_list.extend(err)
                    except ValueError:  # pragma: no cover
                        error_list.append(f"Cannot load data from: {sheet_name}")
        finally:
            xlsx.close()
        data_dict = {}
        for calc_plan in data_list:
            if calc_plan.is_bad():
                error_list.append(f"Problem with load {calc_plan.name} because of {calc_plan.get_error_source()}")
                continue
            new_name = iterate_names(calc_plan.name, data_dict)
            if new_name is None:  # pragma: no cover
                error_list.append(f"Cannot determine proper name for {calc_plan.name}")
            calc_plan.name = new_name
            data_dict[new_name] = calc_plan
        return data_dict, error_list

    @classmethod
    def get_name(cls) -> str:
        return "Calculation plans from result (*.xlsx)"


IO_LABELS_COLORMAP = "io.labels_colormap_dir"
IO_MASK_METADATA_FILE = "metadata.json"
